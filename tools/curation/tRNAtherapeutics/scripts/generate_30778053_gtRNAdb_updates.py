import csv
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


csv.field_size_limit(sys.maxsize)

BASE = Path(__file__).resolve().parents[4] / "field-curation-workdir"
FULL = BASE / "full_tRNAtherapeutics"
BACKUPS = FULL / "backups"
SUPP = FULL / "supplementary" / "30778053" / "41467_2019_8329_MOESM5_ESM.xlsx"
GTRNA = FULL / "external_sources" / "hg19_tRNAs_bundle"
OUT_TSV = FULL / "extractions" / "30778053_gtrnadb_exact_updates.tsv"
OUT_UNRESOLVED = FULL / "extractions" / "30778053_unresolved_origin_anticodon.tsv"
OUT_SQL = FULL / "sql" / "30778053_gtrnadb_exact_updates.sql"
OUT_REPORT = FULL / "reports" / "30778053_gtrnadb_exact_update_review.md"

STOP_TO_ANTI_DNA = {"TGA": "TCA", "TAG": "CTA", "TAA": "TTA"}
PREFIX_RE = re.compile(r"^\s*([A-Za-z]{3})(TGA|TAG|TAA)")
AA_FIELD_RE = re.compile(r"^([A-Za-z]+)\(([AUCG]+)\)$")

GENETIC_CODE = {
    "TTT": "Phe", "TTC": "Phe", "TTA": "Leu", "TTG": "Leu",
    "TCT": "Ser", "TCC": "Ser", "TCA": "Ser", "TCG": "Ser",
    "TAT": "Tyr", "TAC": "Tyr", "TAA": "Stop", "TAG": "Stop",
    "TGT": "Cys", "TGC": "Cys", "TGA": "Stop", "TGG": "Trp",
    "CTT": "Leu", "CTC": "Leu", "CTA": "Leu", "CTG": "Leu",
    "CCT": "Pro", "CCC": "Pro", "CCA": "Pro", "CCG": "Pro",
    "CAT": "His", "CAC": "His", "CAA": "Gln", "CAG": "Gln",
    "CGT": "Arg", "CGC": "Arg", "CGA": "Arg", "CGG": "Arg",
    "ATT": "Ile", "ATC": "Ile", "ATA": "Ile", "ATG": "Met",
    "ACT": "Thr", "ACC": "Thr", "ACA": "Thr", "ACG": "Thr",
    "AAT": "Asn", "AAC": "Asn", "AAA": "Lys", "AAG": "Lys",
    "AGT": "Ser", "AGC": "Ser", "AGA": "Arg", "AGG": "Arg",
    "GTT": "Val", "GTC": "Val", "GTA": "Val", "GTG": "Val",
    "GCT": "Ala", "GCC": "Ala", "GCA": "Ala", "GCG": "Ala",
    "GAT": "Asp", "GAC": "Asp", "GAA": "Glu", "GAG": "Glu",
    "GGT": "Gly", "GGC": "Gly", "GGA": "Gly", "GGG": "Gly",
}


def norm_rna(seq: str) -> str:
    return (seq or "").strip().upper().replace("T", "U").replace(" ", "").replace("\n", "")


def norm_dna(seq: str) -> str:
    return norm_rna(seq).replace("U", "T")


def revcomp_dna(seq: str) -> str:
    return seq.translate(str.maketrans("ATCG", "TAGC"))[::-1]


def qident(name: str) -> str:
    return "`" + name.replace("`", "``") + "`"


def qval(value):
    if value is None or value == "":
        return "NULL"
    return "'" + str(value).replace("\\", "\\\\").replace("'", "''") + "'"


def load_latest_backup():
    paths = sorted(BACKUPS.glob("Engineered_sup_tRNA_PMID30778053_before_*.tsv"))
    if not paths:
        raise RuntimeError("No PMID 30778053 backup TSV found")
    with paths[-1].open(newline="") as handle:
        return paths[-1], list(csv.DictReader(handle, delimiter="\t"))


def load_ace_rows():
    workbook = load_workbook(SUPP, read_only=True, data_only=True)
    sheet = workbook.active
    by_seq = defaultdict(list)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[1] or not row[2]:
            continue
        xlsx_id = str(row[1]).strip()
        seq_raw = str(row[2]).strip()
        match = PREFIX_RE.match(xlsx_id)
        group = None
        if match:
            aa, stop = match.groups()
            group = (aa, stop, STOP_TO_ANTI_DNA[stop].replace("T", "U"))
        item = {
            "xlsx_number": str(row[0]).strip() if row[0] is not None else "",
            "xlsx_id": xlsx_id,
            "seq_raw": seq_raw,
            "seq_dna": norm_dna(seq_raw),
            "group": group,
        }
        by_seq[item["seq_dna"]].append(item)
    return by_seq


def load_fasta_records():
    records_by_seq = defaultdict(list)
    fasta_files = [
        ("hg19_confidence", GTRNA / "hg19-tRNAs.fa"),
        ("hg19_filtered", GTRNA / "hg19-filtered-tRNAs.fa"),
    ]
    for source, path in fasta_files:
        header = None
        seq_parts = []
        for line in path.read_text().splitlines():
            line = line.strip()
            if not line:
                continue
            if line.startswith(">"):
                if header is not None:
                    add_fasta_record(records_by_seq, source, header, "".join(seq_parts))
                header = line[1:]
                seq_parts = []
            else:
                seq_parts.append(line)
        if header is not None:
            add_fasta_record(records_by_seq, source, header, "".join(seq_parts))
    return records_by_seq


def add_fasta_record(records_by_seq, source, header, seq):
    match = re.search(r"tRNAscan-SE ID: ([^)]+)\).*?\b(\w+) \((\w+)\)", header)
    if not match:
        raise RuntimeError(f"Could not parse FASTA header: {header}")
    trnascan_id, aa, anti = match.groups()
    records_by_seq[norm_dna(seq)].append(
        {
            "source": source,
            "header": header,
            "trnascan_id": trnascan_id,
            "aa": aa,
            "anti_dna": anti,
            "seq_dna": norm_dna(seq),
        }
    )


def load_ss_records():
    ss_records = {}
    for path in [GTRNA / "hg19-tRNAs-detailed.ss", GTRNA / "hg19-tRNAs-confidence-set.ss"]:
        current = None
        for line in path.read_text().splitlines():
            id_match = re.match(r"^(chr\S+)\s+\(", line)
            if id_match:
                current = id_match.group(1)
                ss_records.setdefault(current, {})
                continue
            if current and line.startswith("Seq:"):
                ss_records[current]["seq_dna"] = norm_dna(line.split(":", 1)[1])
            elif current and line.startswith("Str:"):
                raw = line.split(":", 1)[1].strip()
                ss_records[current]["structure"] = raw.replace(">", "(").replace("<", ")")
    return ss_records


def candidate_anticodons():
    out = defaultdict(list)
    for codon, aa in GENETIC_CODE.items():
        if aa != "Stop":
            out[aa].append(revcomp_dna(codon))
    return {aa: sorted(set(values)) for aa, values in out.items()}


def match_ace_row(db_row, ace_by_seq):
    aa_match = AA_FIELD_RE.match(db_row.get("aa_and_anticodon_of_sup-tRNA", "") or "")
    aa = aa_match.group(1)[:3] if aa_match else ""
    anti = aa_match.group(2) if aa_match else ""
    candidates = ace_by_seq.get(norm_dna(db_row.get("Sequence_of_sup-tRNA", "")), [])
    filtered = [
        item for item in candidates
        if item["group"] and item["group"][0] == aa and item["group"][2] == anti
    ]
    return filtered or candidates


def classify(base, sup_base):
    if base == "-" and sup_base == "-":
        return "gap"
    if base == "-":
        return "insertion"
    if sup_base == "-":
        return "deletion"
    if base == sup_base:
        return "match"
    return "mismatch"


def decode_json(value):
    return json.loads(value) if value and value != "NULL" else []


def template_spec_from_items(items):
    return [(item["id"], item.get("base", "-") == "-") for item in items]


def build_origin_json(seq, spec):
    seq = norm_rna(seq)
    pos = 0
    out = []
    for coord, is_gap in spec:
        if is_gap:
            base = "-"
        else:
            if pos >= len(seq):
                raise RuntimeError("template consumed beyond sequence")
            base = seq[pos]
            pos += 1
        out.append({"id": coord, "base": base})
    if pos != len(seq):
        raise RuntimeError(f"template consumed {pos}/{len(seq)} bases")
    return out


def build_sup_json(origin_seq, sup_seq, spec):
    origin_seq = norm_rna(origin_seq)
    sup_seq = norm_rna(sup_seq)
    origin_pos = 0
    sup_pos = 0
    out = []
    for coord, is_gap in spec:
        if is_gap:
            base = "-"
            sup_base = "-"
        else:
            if origin_pos >= len(origin_seq) or sup_pos >= len(sup_seq):
                raise RuntimeError("template consumed beyond sequence")
            base = origin_seq[origin_pos]
            sup_base = sup_seq[sup_pos]
            origin_pos += 1
            sup_pos += 1
        out.append({"id": coord, "base": base, "sup_base": sup_base, "type": classify(base, sup_base)})
    if origin_pos != len(origin_seq) or sup_pos != len(sup_seq):
        raise RuntimeError(
            f"template consumed origin {origin_pos}/{len(origin_seq)}, sup {sup_pos}/{len(sup_seq)}"
        )
    return out


def build_template_alignments(updates, db_rows):
    templates = []
    for row in db_rows:
        if not norm_rna(row.get("Sequence_of_origin_tRNA")):
            continue
        if not row.get("js_origin_tRNA") or row["js_origin_tRNA"] == "[]":
            continue
        try:
            items = decode_json(row["js_origin_tRNA"])
            spec = template_spec_from_items(items)
        except Exception:
            continue
        non_gap = sum(1 for _, is_gap in spec if not is_gap)
        templates.append(
            {
                "ENSURE_ID": row["ENSURE_ID"],
                "origin_aa": row.get("aa_and_anticodon_of_origin_tRNA", ""),
                "origin_aa_name": (row.get("aa_and_anticodon_of_origin_tRNA", "").split("(", 1)[0]),
                "origin_len": len(norm_rna(row.get("Sequence_of_origin_tRNA", ""))),
                "non_gap": non_gap,
                "spec": spec,
            }
        )

    for row in updates:
        origin_len = len(norm_rna(row["Sequence_of_origin_tRNA"]))
        origin_aa = row["aa_and_anticodon_of_origin_tRNA"]
        origin_aa_name = origin_aa.split("(", 1)[0]
        candidates = [tpl for tpl in templates if tpl["non_gap"] == origin_len]
        exact = [tpl for tpl in candidates if tpl["origin_aa"] == origin_aa]
        same_aa = [tpl for tpl in candidates if tpl["origin_aa_name"] == origin_aa_name]
        chosen = (exact or same_aa or candidates)
        if not chosen:
            row["alignment_method"] = "not_generated_no_template"
            row["sprinzl_template_id"] = ""
            continue
        template = chosen[0]
        js_origin = build_origin_json(row["Sequence_of_origin_tRNA"], template["spec"])
        js_sup = build_sup_json(row["Sequence_of_origin_tRNA"], row["Sequence_of_sup-tRNA"], template["spec"])
        row["coord_count"] = str(len(js_origin))
        row["js_origin_tRNA"] = json.dumps(js_origin, ensure_ascii=False)
        row["js_sup_tRNA"] = json.dumps(js_sup, ensure_ascii=False)
        row["alignment_method"] = "existing_db_sprinzl_template"
        row["sprinzl_template_id"] = template["ENSURE_ID"]


def main():
    backup_path, db_rows = load_latest_backup()
    ace_by_seq = load_ace_rows()
    fasta_by_seq = load_fasta_records()
    ss_by_id = load_ss_records()
    anticodons_by_aa = candidate_anticodons()

    updates = []
    unresolved = []
    for db_row in db_rows:
        if norm_rna(db_row.get("Sequence_of_origin_tRNA")):
            continue

        ace_matches = match_ace_row(db_row, ace_by_seq)
        if len(ace_matches) != 1 or not ace_matches[0].get("group"):
            unresolved.append({
                "ENSURE_ID": db_row["ENSURE_ID"],
                "Index": db_row["Index"],
                "aa_and_anticodon_of_sup-tRNA": db_row["aa_and_anticodon_of_sup-tRNA"],
                "reason": f"ACE row match count={len(ace_matches)} or missing group",
                "xlsx_id": ";".join(item.get("xlsx_id", "") for item in ace_matches),
            })
            continue

        ace = ace_matches[0]
        aa, stop, _sup_anti = ace["group"]
        stop_anti_dna = STOP_TO_ANTI_DNA[stop]
        stop_matches = list(re.finditer(stop_anti_dna.lower(), ace["seq_raw"]))
        if len(stop_matches) != 1:
            unresolved.append({
                "ENSURE_ID": db_row["ENSURE_ID"],
                "Index": db_row["Index"],
                "aa_and_anticodon_of_sup-tRNA": db_row["aa_and_anticodon_of_sup-tRNA"],
                "reason": f"red/lowcase stop anticodon occurrence count={len(stop_matches)}",
                "xlsx_id": ace["xlsx_id"],
            })
            continue
        pos = stop_matches[0].start()

        hits = []
        for candidate in anticodons_by_aa[aa]:
            parent_dna = (ace["seq_raw"][:pos] + candidate + ace["seq_raw"][pos + 3:]).upper()
            for record in fasta_by_seq.get(norm_dna(parent_dna), []):
                if record["aa"] == aa and record["anti_dna"] == candidate:
                    hits.append((candidate, parent_dna, record))

        if len(hits) != 1:
            unresolved.append({
                "ENSURE_ID": db_row["ENSURE_ID"],
                "Index": db_row["Index"],
                "aa_and_anticodon_of_sup-tRNA": db_row["aa_and_anticodon_of_sup-tRNA"],
                "reason": f"GtRNAdb exact candidate hits={len(hits)}",
                "xlsx_id": ace["xlsx_id"],
            })
            continue

        anti_dna, parent_dna, record = hits[0]
        ss = ss_by_id.get(record["trnascan_id"])
        if not ss or ss.get("seq_dna") != norm_dna(parent_dna):
            unresolved.append({
                "ENSURE_ID": db_row["ENSURE_ID"],
                "Index": db_row["Index"],
                "aa_and_anticodon_of_sup-tRNA": db_row["aa_and_anticodon_of_sup-tRNA"],
                "reason": f"missing/mismatched GtRNAdb secondary structure for {record['trnascan_id']}",
                "xlsx_id": ace["xlsx_id"],
            })
            continue
        structure = ss["structure"]
        origin_rna = norm_rna(parent_dna)
        if len(structure) != len(origin_rna):
            raise RuntimeError(f"{db_row['ENSURE_ID']}: structure length mismatch")

        updates.append({
            "ENSURE_ID": db_row["ENSURE_ID"],
            "Index": db_row["Index"],
            "xlsx_id": ace["xlsx_id"],
            "xlsx_number": ace["xlsx_number"],
            "matched_gtrnadb_source": record["source"],
            "tRNAscan-SE_ID_of_origin_tRNA": record["trnascan_id"],
            "aa_and_anticodon_of_origin_tRNA": f"{aa}({anti_dna.replace('T', 'U')})",
            "Sequence_of_origin_tRNA": origin_rna,
            "Sequence_of_sup-tRNA": norm_rna(db_row["Sequence_of_sup-tRNA"]),
            "Secondary structure": structure,
            "Secondary structure of sup-trna": db_row["Secondary structure of sup-trna"],
            "current_pdbid": db_row["pdbid"],
            "evidence": "PMID 30778053 Supplementary Data 1 ACE sequence; GtRNAdb hg19 exact parent-sequence match; GtRNAdb detailed.ss structure",
            "js_origin_tRNA": "",
            "js_sup_tRNA": "",
            "coord_count": "",
            "alignment_method": "",
            "sprinzl_template_id": "",
        })

    build_template_alignments(updates, db_rows)

    OUT_TSV.parent.mkdir(parents=True, exist_ok=True)
    with OUT_TSV.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(updates[0].keys()), delimiter="\t")
        writer.writeheader()
        writer.writerows(updates)

    with OUT_UNRESOLVED.open("w", newline="") as handle:
        fieldnames = ["ENSURE_ID", "Index", "aa_and_anticodon_of_sup-tRNA", "xlsx_id", "reason"]
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        writer.writerows(unresolved)

    sql = [
        "-- PMID 30778053 exact GtRNAdb-backed origin/Sprinzl updates.",
        f"-- DB backup: {backup_path}",
        f"-- Review TSV: {OUT_TSV}",
        "START TRANSACTION;",
    ]
    for row in updates:
        assignments = [
            f"{qident('aa_and_anticodon_of_origin_tRNA')} = {qval(row['aa_and_anticodon_of_origin_tRNA'])}",
            f"{qident('tRNAscan-SE_ID_of_origin_tRNA')} = {qval(row['tRNAscan-SE_ID_of_origin_tRNA'])}",
            f"{qident('Sequence_of_origin_tRNA')} = {qval(row['Sequence_of_origin_tRNA'])}",
            f"{qident('Secondary structure')} = {qval(row['Secondary structure'])}",
        ]
        if row["js_origin_tRNA"] and row["js_sup_tRNA"]:
            assignments.extend([
                f"{qident('js_origin_tRNA')} = {qval(row['js_origin_tRNA'])}",
                f"{qident('js_sup_tRNA')} = {qval(row['js_sup_tRNA'])}",
            ])
        sql.append(
            "\nUPDATE Engineered_sup_tRNA\nSET\n  "
            + ",\n  ".join(assignments)
            + f"\nWHERE PMID = 30778053 AND ENSURE_ID = {qval(row['ENSURE_ID'])};"
        )
    sql.append("\nCOMMIT;")
    OUT_SQL.parent.mkdir(parents=True, exist_ok=True)
    OUT_SQL.write_text("\n".join(sql) + "\n")

    max_origin = max(len(row["js_origin_tRNA"]) for row in updates)
    max_sup = max(len(row["js_sup_tRNA"]) for row in updates)
    with_alignment = sum(1 for row in updates if row["js_origin_tRNA"] and row["js_sup_tRNA"])
    OUT_REPORT.parent.mkdir(parents=True, exist_ok=True)
    OUT_REPORT.write_text(
        "\n".join([
            "# PMID 30778053 exact GtRNAdb-backed updates",
            "",
            "This review file includes only rows where the missing parent/origin tRNA can be reconstructed without choosing among ambiguous anticodons.",
            "",
            "Evidence chain:",
            "",
            "1. The article methods state that tRNA gene sequences were obtained from tRNAscan-SE/GtRNAdb and that anticodons were mutated to UAG/UGA/UAA.",
            "2. Supplementary Data 1 (`41467_2019_8329_MOESM5_ESM.xlsx`) provides the ACE-tRNA sequences with the engineered suppressor anticodon highlighted in red/lowercase.",
            "3. For each accepted row, replacing only that suppressor anticodon with one candidate parent anticodon produced exactly one hg19 GtRNAdb confidence/filtered sequence with matching amino acid and anticodon.",
            "4. The origin secondary structure is copied from the matching GtRNAdb `hg19-tRNAs-detailed.ss` entry.",
            "5. `js_origin_tRNA` and `js_sup_tRNA` were generated from existing database Sprinzl templates with the same non-gap sequence length, preferring the same origin amino acid/anticodon. Sequence preservation is deterministic because bases are consumed left-to-right into the template coordinates.",
            "",
            f"Accepted updates: {len(updates)}",
            f"Accepted updates with Sprinzl/alignment JSON: {with_alignment}",
            f"Unresolved/ambiguous rows left untouched: {len(unresolved)}",
            f"Max `js_origin_tRNA` length: {max_origin}",
            f"Max `js_sup_tRNA` length: {max_sup}",
            "",
            f"Review TSV: `{OUT_TSV}`",
            f"Unresolved TSV: `{OUT_UNRESOLVED}`",
            f"SQL: `{OUT_SQL}`",
        ]) + "\n"
    )

    print(OUT_TSV)
    print(OUT_UNRESOLVED)
    print(OUT_SQL)
    print(OUT_REPORT)
    print("accepted", len(updates))
    print("unresolved", len(unresolved))
    print("with_alignment", with_alignment)
    print("max_origin_json", max_origin)
    print("max_sup_json", max_sup)


if __name__ == "__main__":
    main()
