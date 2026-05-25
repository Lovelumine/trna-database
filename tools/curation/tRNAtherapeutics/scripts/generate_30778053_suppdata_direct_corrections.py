#!/usr/bin/env python3
"""Generate direct Supplementary Data 1 corrections for PMID 30778053.

This script only updates fields that are explicitly present in the article's
Supplementary Data 1 sequence table. It does not infer origin sequences.
"""

from __future__ import annotations

import csv
import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


csv.field_size_limit(sys.maxsize)

BASE = Path(__file__).resolve().parents[4] / "field-curation-workdir"
FULL = BASE / "full_tRNAtherapeutics"
BACKUPS = FULL / "backups"
SUPP = FULL / "supplementary" / "30778053" / "41467_2019_8329_MOESM5_ESM.xlsx"
OUT_TSV = FULL / "extractions" / "30778053_suppdata_direct_corrections.tsv"
OUT_UNRESOLVED = FULL / "extractions" / "30778053_suppdata_direct_corrections_skipped.tsv"
OUT_SQL = FULL / "sql" / "30778053_suppdata_direct_corrections.sql"
OUT_REPORT = FULL / "reports" / "30778053_suppdata_direct_correction_review.md"


def norm_rna(seq: str) -> str:
    return (seq or "").strip().upper().replace("T", "U").replace(" ", "").replace("\n", "")


def qident(name: str) -> str:
    return "`" + name.replace("`", "``") + "`"


def qval(value: object) -> str:
    if value is None or value == "":
        return "NULL"
    return "'" + str(value).replace("\\", "\\\\").replace("'", "''") + "'"


def latest_backup() -> Path:
    paths = sorted(BACKUPS.glob("Engineered_sup_tRNA_PMID30778053_before_suppdata_direct_*.tsv"))
    if not paths:
        raise RuntimeError("No suppdata-direct backup found for PMID 30778053")
    return paths[-1]


def load_backup_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def load_ace_rows() -> dict[str, list[dict[str, str]]]:
    workbook = load_workbook(SUPP, read_only=True, data_only=True)
    sheet = workbook.active
    by_seq: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[1] or not row[2]:
            continue
        xlsx_number = str(row[0]).strip() if row[0] is not None else ""
        xlsx_id = str(row[1]).strip()
        seq_raw = str(row[2]).strip()
        by_seq[norm_rna(seq_raw)].append(
            {
                "xlsx_number": xlsx_number,
                "xlsx_id": xlsx_id,
                "seq_raw": seq_raw,
            }
        )
    return by_seq


def extract_single_trnascan_id(xlsx_id: str) -> tuple[str, str]:
    normalized = re.sub(r"c\s+hr", "chr", xlsx_id)
    ids = re.findall(r"chr(?:[0-9XYM]+|Un\d+)\.trna\d+", normalized)
    unique = sorted(set(ids), key=ids.index)
    if len(unique) == 1:
        return unique[0], "single_id"
    if len(unique) > 1:
        return ";".join(unique), "multiple_ids"
    return "", "no_id"


def is_blank(value: str | None) -> bool:
    return value is None or str(value).strip() == "" or str(value).strip().upper() == "NULL"


def main() -> None:
    backup_path = latest_backup()
    rows = load_backup_rows(backup_path)
    ace_by_seq = load_ace_rows()

    updates: list[dict[str, str]] = []
    skipped: list[dict[str, str]] = []
    for row in rows:
        if not is_blank(row.get("tRNAscan-SE_ID_of_origin_tRNA")):
            continue
        sup_seq = norm_rna(row.get("Sequence_of_sup-tRNA", ""))
        if not sup_seq:
            skipped.append(
                {
                    "ENSURE_ID": row["ENSURE_ID"],
                    "Index": row["Index"],
                    "aa_and_anticodon_of_sup-tRNA": row["aa_and_anticodon_of_sup-tRNA"],
                    "xlsx_id": "",
                    "reason": "missing suppressor sequence",
                }
            )
            continue
        ace_matches = ace_by_seq.get(sup_seq, [])
        if len(ace_matches) != 1:
            skipped.append(
                {
                    "ENSURE_ID": row["ENSURE_ID"],
                    "Index": row["Index"],
                    "aa_and_anticodon_of_sup-tRNA": row["aa_and_anticodon_of_sup-tRNA"],
                    "xlsx_id": ";".join(item["xlsx_id"] for item in ace_matches),
                    "reason": f"ACE sequence match count={len(ace_matches)}",
                }
            )
            continue
        ace = ace_matches[0]
        trnascan_id, id_status = extract_single_trnascan_id(ace["xlsx_id"])
        if id_status != "single_id":
            skipped.append(
                {
                    "ENSURE_ID": row["ENSURE_ID"],
                    "Index": row["Index"],
                    "aa_and_anticodon_of_sup-tRNA": row["aa_and_anticodon_of_sup-tRNA"],
                    "xlsx_id": ace["xlsx_id"],
                    "reason": id_status,
                }
            )
            continue

        updates.append(
            {
                "ENSURE_ID": row["ENSURE_ID"],
                "Index": row["Index"],
                "aa_and_anticodon_of_sup-tRNA": row["aa_and_anticodon_of_sup-tRNA"],
                "current_tRNAscan-SE_ID_of_origin_tRNA": row["tRNAscan-SE_ID_of_origin_tRNA"],
                "new_tRNAscan-SE_ID_of_origin_tRNA": trnascan_id,
                "xlsx_number": ace["xlsx_number"],
                "xlsx_id": ace["xlsx_id"],
                "Sequence_of_sup-tRNA": sup_seq,
                "origin_sequence_blank": str(int(is_blank(row.get("Sequence_of_origin_tRNA")))),
                "evidence": (
                    "PMID 30778053 Supplementary Data 1 "
                    "(`41467_2019_8329_MOESM5_ESM.xlsx`, sheet `ACE-tRNA sequences`) "
                    "lists this exact ACE-tRNA sequence and tRNAscan-SE ID."
                ),
                "confidence": "high",
            }
        )

    OUT_TSV.parent.mkdir(parents=True, exist_ok=True)
    with OUT_TSV.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(updates[0].keys()), delimiter="\t")
        writer.writeheader()
        writer.writerows(updates)

    with OUT_UNRESOLVED.open("w", newline="", encoding="utf-8") as handle:
        fieldnames = ["ENSURE_ID", "Index", "aa_and_anticodon_of_sup-tRNA", "xlsx_id", "reason"]
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        writer.writerows(skipped)

    sql = [
        "-- PMID 30778053 direct Supplementary Data 1 tRNAscan-SE corrections.",
        f"-- DB backup: {backup_path}",
        f"-- Review TSV: {OUT_TSV}",
        "START TRANSACTION;",
    ]
    for update in updates:
        sql.append(
            "\nUPDATE Engineered_sup_tRNA\nSET\n  "
            + f"{qident('tRNAscan-SE_ID_of_origin_tRNA')} = {qval(update['new_tRNAscan-SE_ID_of_origin_tRNA'])}"
            + f"\nWHERE PMID = 30778053 AND ENSURE_ID = {qval(update['ENSURE_ID'])};"
        )
    sql.append("\nCOMMIT;")
    OUT_SQL.parent.mkdir(parents=True, exist_ok=True)
    OUT_SQL.write_text("\n".join(sql) + "\n", encoding="utf-8")

    OUT_REPORT.parent.mkdir(parents=True, exist_ok=True)
    OUT_REPORT.write_text(
        "\n".join(
            [
                "# PMID 30778053 Supplementary Data 1 direct corrections",
                "",
                "Scope: fill only `tRNAscan-SE_ID_of_origin_tRNA` values explicitly recoverable from Supplementary Data 1.",
                "",
                f"Accepted updates: {len(updates)}",
                f"Skipped rows: {len(skipped)}",
                "",
                "Evidence chain:",
                "",
                "1. Supplementary Data 1 (`41467_2019_8329_MOESM5_ESM.xlsx`) contains the `ACE-tRNA sequences` sheet with columns `tRNAscan-SE ID` and `ACE-tRNA sequence 5' -> 3'`.",
                "2. Each accepted database row was matched by exact suppressor sequence to exactly one Supplementary Data 1 row.",
                "3. The new value is the single `chr*.trna*` identifier present in that Supplementary Data 1 ID. A minor OCR/spacing artifact `c hr7.trna32` was normalized to `chr7.trna32` only after exact sequence matching.",
                "4. Origin sequence, origin secondary structure, and Sprinzl JSON are intentionally not inferred in this correction batch.",
                "",
                "Skipped rows mostly have either an ambiguous ACE sequence mapping to two IDs, or are outside Supplementary Data 1 species/suppressor rows.",
                "",
                f"Review TSV: `{OUT_TSV}`",
                f"Skipped TSV: `{OUT_UNRESOLVED}`",
                f"SQL: `{OUT_SQL}`",
                f"Backup: `{backup_path}`",
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    print(OUT_TSV)
    print(OUT_UNRESOLVED)
    print(OUT_SQL)
    print(OUT_REPORT)
    print("accepted", len(updates))
    print("skipped", len(skipped))


if __name__ == "__main__":
    main()
